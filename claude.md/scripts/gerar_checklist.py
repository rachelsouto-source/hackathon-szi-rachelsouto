#!/usr/bin/env python3
"""
Gera a planilha de controle da DD Técnica (.xlsx) a partir de um JSON de achados.

Uso:
    python gerar_checklist.py achados.json "Jurere Spot III" saida.xlsx

Formato do JSON de entrada (achados.json):
{
  "empreendimento": "Jurerê Spot III",
  "data": "2025-09-23",
  "itens": [
    {
      "etapa": "Matrícula",
      "documento": "Matrícula 6.698 / 7.568",
      "status": "Divergência",          # OK | Pendente | Divergência | Não se aplica
      "severidade": "Crítico",          # OK | Atenção | Crítico
      "observacao": "Área registrada 1.382,40 m² difere do topográfico (1.158,96 m²).",
      "acao": "Realizar retificação de matrícula e amembramento.",
      "fonte": "Matrícula (ônus e ações).pdf"
    }
  ]
}

Requer: openpyxl  (pip install openpyxl)
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl não instalado. Rode: pip install openpyxl")

# Cores por status
FILL = {
    "OK": "C6EFCE",
    "Pendente": "FFEB9C",
    "Divergência": "FFC7CE",
    "Divergencia": "FFC7CE",
    "Não se aplica": "D9D9D9",
    "Nao se aplica": "D9D9D9",
}
FONT_COLOR = {
    "OK": "006100",
    "Pendente": "9C6500",
    "Divergência": "9C0006",
    "Divergencia": "9C0006",
    "Não se aplica": "595959",
    "Nao se aplica": "595959",
}
HEADERS = ["#", "Etapa", "Documento", "Status", "Severidade", "Observação", "Ação recomendada", "Fonte"]


def gerar(achados_path: str, empreendimento: str, saida: str) -> None:
    data = json.loads(Path(achados_path).read_text(encoding="utf-8"))
    itens = data.get("itens", [])
    nome = data.get("empreendimento", empreendimento)

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle DD Técnica"

    # Título
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Controle DD Técnica — {nome}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Cabeçalho
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Linhas
    for i, item in enumerate(itens, start=1):
        r = i + 2
        status = item.get("status", "Pendente")
        row_vals = [
            i,
            item.get("etapa", ""),
            item.get("documento", ""),
            status,
            item.get("severidade", ""),
            item.get("observacao", ""),
            item.get("acao", ""),
            item.get("fonte", ""),
        ]
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        # Colorir a célula de status
        sc = ws.cell(row=r, column=4)
        if status in FILL:
            sc.fill = PatternFill("solid", fgColor=FILL[status])
            sc.font = Font(bold=True, color=FONT_COLOR.get(status, "000000"))
            sc.alignment = Alignment(horizontal="center", vertical="center")

    # Larguras
    widths = [4, 22, 26, 14, 12, 48, 40, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    # Aba resumo
    ws2 = wb.create_sheet("Resumo")
    from collections import Counter
    cnt = Counter(it.get("status", "Pendente") for it in itens)
    ws2["A1"] = "Resumo por status"
    ws2["A1"].font = Font(bold=True, size=12)
    row = 3
    for status in ["OK", "Pendente", "Divergência", "Não se aplica"]:
        ws2.cell(row=row, column=1, value=status)
        cell = ws2.cell(row=row, column=2, value=cnt.get(status, 0))
        if status in FILL:
            ws2.cell(row=row, column=1).fill = PatternFill("solid", fgColor=FILL[status])
        row += 1
    ws2.cell(row=row + 1, column=1, value="Total de itens").font = Font(bold=True)
    ws2.cell(row=row + 1, column=2, value=len(itens)).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10

    wb.save(saida)
    print(f"Planilha gerada: {saida}  ({len(itens)} itens)")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.exit("Uso: python gerar_checklist.py <achados.json> <empreendimento> <saida.xlsx>")
    gerar(sys.argv[1], sys.argv[2], sys.argv[3])
